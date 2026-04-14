import os
import json
import time
import base64
import logging
import zipfile
import io
import csv
from datetime import datetime
from typing import Dict, Optional, List, Any

import boto3
import requests
from PyPDF2 import PdfReader

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


logger = logging.getLogger()
logger.setLevel(logging.INFO)


DOC2ORDER_BUCKET = os.environ["DOC2ORDER_BUCKET"]
QUARANTINE_PREFIX = os.environ["QUARANTINE_PREFIX"].rstrip("/") + "/"
STATE_KEY = os.environ["STATE_KEY"]
GRAPH_SECRET_ARN = os.environ["GRAPH_SECRET_ARN"]

APP_ENV = os.environ.get("APP_ENV", "dev")
DEFAULT_CUSTOMER_ID = os.environ.get("CUSTOMER_ID", "evapo")

MAX_MESSAGES_PER_RUN = int(os.environ.get("MAX_MESSAGES_PER_RUN", "25"))
PROCESSED_FOLDER_NAME = os.environ.get("PROCESSED_FOLDER_NAME", "Processed/Doc2OrderInquiryEngine")

MAX_PDF_TEXT_BYTES = int(os.environ.get("MAX_PDF_TEXT_BYTES", "500000"))
MAX_PDF_SIZE_MB = int(os.environ.get("MAX_PDF_SIZE_MB", "15"))
STORE_EMAIL_METADATA = os.environ.get("STORE_EMAIL_METADATA", "true").lower() == "true"

MAX_PDF_BYTES = MAX_PDF_SIZE_MB * 1024 * 1024


def _safe_json_loads(s: str):
    try:
        return json.loads(s)
    except Exception:
        return None


INQUIRY_KEYWORDS_JSON = os.environ.get("INQUIRY_KEYWORDS_JSON", "").strip()
_inquiry_keywords = _safe_json_loads(INQUIRY_KEYWORDS_JSON) if INQUIRY_KEYWORDS_JSON else None
if not isinstance(_inquiry_keywords, list) or not _inquiry_keywords:
    _inquiry_keywords = [
        "stock inquiry",
        "stock enquiry",
        "inventory inquiry",
        "inventory enquiry",
        "availability inquiry",
        "availability enquiry",
        "inquiry",
        "enquiry",
    ]

INQUIRY_KEYWORDS = [str(x).strip().lower() for x in _inquiry_keywords if str(x).strip()]

CUSTOMER_ROUTING_RULES_JSON = os.environ.get("CUSTOMER_ROUTING_RULES_JSON", "").strip()
_customer_rules = _safe_json_loads(CUSTOMER_ROUTING_RULES_JSON) if CUSTOMER_ROUTING_RULES_JSON else None
if _customer_rules is None:
    _customer_rules = []

CUSTOMER_ROUTING_RULES: List[Dict[str, Any]] = []
for r in _customer_rules:
    if not isinstance(r, dict):
        continue
    cid = (r.get("customer_id") or "").strip()
    if not cid:
        continue
    CUSTOMER_ROUTING_RULES.append({
        "customer_id": cid.lower(),
        "keywords": [str(k).strip().lower() for k in (r.get("keywords") or []) if str(k).strip()],
        "from_domains": [str(d).strip().lower() for d in (r.get("from_domains") or []) if str(d).strip()],
        "from_emails": [str(e).strip().lower() for e in (r.get("from_emails") or []) if str(e).strip()],
    })

if not CUSTOMER_ROUTING_RULES:
    CUSTOMER_ROUTING_RULES = [
        {
            "customer_id": "evapo",
            "keywords": ["evapo", "spatial global", "evapo c/o spatial"],
            "from_domains": [],
            "from_emails": [],
        }
    ]


s3 = boto3.client("s3")
secretsmanager = boto3.client("secretsmanager")

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
TOKEN_URL_TEMPLATE = "https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"

SUPPORTED_EXTENSIONS = (".pdf", ".csv", ".xlsx", ".xlsm")


def is_supported_doc(filename: str) -> bool:
    return (filename or "").lower().endswith(SUPPORTED_EXTENSIONS)


def get_mail_folder_id(mailbox: str, folder_path: str, token: str) -> str:
    parts = [p.strip() for p in folder_path.split("/") if p.strip()]
    if not parts:
        raise ValueError("PROCESSED_FOLDER_NAME is empty")

    parent_id = None

    for part in parts:
        if parent_id is None:
            list_url = f"{GRAPH_BASE}/users/{mailbox}/mailFolders"
            create_url = list_url
        else:
            list_url = f"{GRAPH_BASE}/users/{mailbox}/mailFolders/{parent_id}/childFolders"
            create_url = list_url

        folders = graph_get(list_url, token).get("value", [])
        match = next((f for f in folders if f.get("displayName") == part), None)

        if match:
            parent_id = match["id"]
        else:
            created = graph_post(create_url, token, {"displayName": part})
            parent_id = created["id"]

    return parent_id


def load_graph_secret() -> Dict[str, str]:
    resp = secretsmanager.get_secret_value(SecretId=GRAPH_SECRET_ARN)
    return json.loads(resp["SecretString"])


def get_graph_token(secret: Dict[str, str]) -> str:
    url = TOKEN_URL_TEMPLATE.format(tenant_id=secret["TENANT_ID"])
    r = requests.post(url, data={
        "client_id": secret["CLIENT_ID"],
        "client_secret": secret["CLIENT_SECRET"],
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def graph_get(url: str, token: str) -> Dict:
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    if not r.ok:
        try:
            err = r.json()
        except Exception:
            err = {"raw_text": r.text}
        logger.error("Graph GET failed: status=%s url=%s body=%s", r.status_code, url, json.dumps(err)[:2000])
        r.raise_for_status()
    return r.json()


def graph_post(url: str, token: str, payload: Dict) -> Dict:
    r = requests.post(
        url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json=payload,
        timeout=30,
    )
    if not r.ok:
        try:
            err = r.json()
        except Exception:
            err = {"raw_text": r.text}
        logger.error("Graph POST failed: status=%s url=%s body=%s", r.status_code, url, json.dumps(err)[:2000])
        r.raise_for_status()
    return r.json() if r.text else {}


def graph_move_message(mailbox: str, msg_id: str, folder_id: str, token: str) -> None:
    graph_post(
        f"{GRAPH_BASE}/users/{mailbox}/messages/{msg_id}/move",
        token,
        {"destinationId": folder_id},
    )


def s3_put_bytes(key: str, content: bytes) -> None:
    s3.put_object(Bucket=DOC2ORDER_BUCKET, Key=key, Body=content)


def extract_pdf_text(file_bytes: bytes) -> str:
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
            if len(text.encode("utf-8")) > MAX_PDF_TEXT_BYTES:
                break
        return text.lower()
    except Exception:
        return ""


def extract_csv_text(file_bytes: bytes) -> str:
    try:
        decoded = file_bytes.decode("utf-8", errors="replace")
        out_parts = []
        reader = csv.reader(io.StringIO(decoded))
        for row in reader:
            if row:
                out_parts.append(" ".join([str(x) for x in row if x is not None]))
            if len(" ".join(out_parts).encode("utf-8")) > MAX_PDF_TEXT_BYTES:
                break
        return "\n".join(out_parts).lower()
    except Exception:
        return ""


def extract_excel_text(file_bytes: bytes) -> str:
    if load_workbook is None:
        logger.warning("openpyxl not available; Excel text extraction skipped.")
        return ""

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        out_parts = []

        for ws in wb.worksheets:
            out_parts.append(ws.title)
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) for v in row if v is not None and str(v).strip()]
                if vals:
                    out_parts.append(" ".join(vals))
                if len("\n".join(out_parts).encode("utf-8")) > MAX_PDF_TEXT_BYTES:
                    return "\n".join(out_parts).lower()

        return "\n".join(out_parts).lower()
    except Exception:
        return ""


def extract_attachment_text(filename: str, file_bytes: bytes) -> str:
    filename_l = (filename or "").lower()

    if filename_l.endswith(".pdf"):
        return extract_pdf_text(file_bytes)

    if filename_l.endswith(".csv"):
        return extract_csv_text(file_bytes)

    if filename_l.endswith(".xlsx") or filename_l.endswith(".xlsm"):
        return extract_excel_text(file_bytes)

    return ""


def detect_inquiry_intent(subject: str, body: str, filename: str, file_text: Optional[str]) -> Dict[str, Any]:
    subject_l = (subject or "").lower()
    body_l = (body or "").lower()
    filename_l = (filename or "").lower()
    file_l = file_text or ""

    haystacks = [subject_l, body_l, filename_l, file_l]
    matched = sorted({kw for kw in INQUIRY_KEYWORDS for h in haystacks if kw in h})

    return {
        "is_inquiry": len(matched) > 0,
        "matched_keywords": matched,
    }


def detect_customer_with_score(
    subject: str,
    body: str,
    filename: str,
    file_text: Optional[str],
    sender_email: str,
) -> Optional[Dict[str, Any]]:
    subject_l = (subject or "").lower()
    body_l = (body or "").lower()
    filename_l = (filename or "").lower()
    file_l = file_text or ""
    sender_l = (sender_email or "").lower().strip()
    sender_domain = sender_l.split("@")[-1] if "@" in sender_l else ""

    haystacks = [subject_l, body_l, filename_l, file_l]

    best: Optional[Dict[str, Any]] = None

    for rule in CUSTOMER_ROUTING_RULES:
        score = 0
        basis: List[str] = []
        matched_keywords: List[str] = []

        if sender_l and sender_l in rule["from_emails"]:
            score += 100
            basis.append(f"from_email:{sender_l}")

        if sender_domain and sender_domain in rule["from_domains"]:
            score += 70
            basis.append(f"from_domain:{sender_domain}")

        for kw in rule["keywords"]:
            for h in haystacks:
                if kw and kw in h:
                    score += 30
                    matched_keywords.append(kw)
                    basis.append(f"keyword:{kw}")
                    break

        matched_keywords = sorted(list(set(matched_keywords)))
        basis = sorted(list(set(basis)))

        if score <= 0:
            continue

        candidate = {
            "customer_id": rule["customer_id"],
            "score": score,
            "basis": basis,
            "matched_keywords": matched_keywords,
        }

        if best is None or candidate["score"] > best["score"]:
            best = candidate

    return best


def load_state() -> Dict:
    try:
        obj = s3.get_object(Bucket=DOC2ORDER_BUCKET, Key=STATE_KEY)
        return json.loads(obj["Body"].read())
    except s3.exceptions.NoSuchKey:
        return {"processed_message_ids": []}


def save_state(state: Dict) -> None:
    s3.put_object(
        Bucket=DOC2ORDER_BUCKET,
        Key=STATE_KEY,
        Body=json.dumps(state).encode("utf-8"),
    )


def lambda_handler(event, context):
    secret = load_graph_secret()
    token = get_graph_token(secret)
    mailbox = secret["MAILBOX_UPN"]

    processed_folder_id = get_mail_folder_id(mailbox, PROCESSED_FOLDER_NAME, token)

    state = load_state()
    processed_ids = set(state.get("processed_message_ids", []))
    newly_processed = []

    messages = graph_get(
        f"{GRAPH_BASE}/users/{mailbox}/mailFolders/Inbox/messages"
        f"?$filter=isRead eq false&$top={MAX_MESSAGES_PER_RUN}",
        token,
    ).get("value", [])

    for msg in messages:
        msg_id = msg["id"]
        if msg_id in processed_ids:
            continue

        subject = msg.get("subject", "")
        body = msg.get("body", {}).get("content", "")
        sender = msg.get("from", {}).get("emailAddress", {}).get("address", "")

        attachments = graph_get(
            f"{GRAPH_BASE}/users/{mailbox}/messages/{msg_id}/attachments",
            token,
        ).get("value", [])

        any_attachment_processed = False
        inquiry_detected_for_message = False

        for att in attachments:
            if not att.get("contentBytes"):
                continue

            raw_bytes = base64.b64decode(att["contentBytes"])
            files = []

            if att["name"].lower().endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(raw_bytes)) as z:
                    for zi in z.infolist():
                        if is_supported_doc(zi.filename):
                            files.append((zi.filename, z.read(zi)))
            elif is_supported_doc(att["name"]):
                files.append((att["name"], raw_bytes))

            for filename, file_bytes in files:
                ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
                safe_name = filename.replace(" ", "_")

                metadata = {
                    "message_id": msg_id,
                    "sender": sender,
                    "subject": subject,
                    "original_filename": filename,
                    "received_at": msg.get("receivedDateTime"),
                }

                if len(file_bytes) > MAX_PDF_BYTES:
                    metadata["rejected_reason"] = "file_too_large"
                    key = f"{QUARANTINE_PREFIX}{ts}_{msg_id}_{safe_name}"
                else:
                    file_text = extract_attachment_text(filename, file_bytes)

                    inquiry_result = detect_inquiry_intent(subject, body, filename, file_text)
                    routing_result = detect_customer_with_score(subject, body, filename, file_text, sender)

                    metadata["inquiry_detected"] = inquiry_result["is_inquiry"]
                    metadata["inquiry_keywords"] = inquiry_result["matched_keywords"]

                    if inquiry_result["is_inquiry"]:
                        inquiry_detected_for_message = True

                    if not inquiry_result["is_inquiry"]:
                        metadata["rejected_reason"] = "not_inquiry_email"
                        key = f"{QUARANTINE_PREFIX}{ts}_{msg_id}_{safe_name}"
                    elif not routing_result:
                        metadata["rejected_reason"] = "customer_not_detected"
                        key = f"{QUARANTINE_PREFIX}{ts}_{msg_id}_{safe_name}"
                    else:
                        customer_id = routing_result["customer_id"]

                        metadata["customer_id"] = customer_id
                        metadata["routing_score"] = routing_result["score"]
                        metadata["routing_basis"] = routing_result["basis"]
                        metadata["routing_matched_keywords"] = routing_result["matched_keywords"]

                        key = f"input/{APP_ENV}/{customer_id}/{ts}_{msg_id}_{safe_name}"
                        any_attachment_processed = True

                s3_put_bytes(key, file_bytes)

                if STORE_EMAIL_METADATA:
                    s3_put_bytes(key + ".metadata.json", json.dumps(metadata).encode("utf-8"))

        if any_attachment_processed or attachments or inquiry_detected_for_message:
            graph_move_message(mailbox, msg_id, processed_folder_id, token)

        newly_processed.append(msg_id)
        time.sleep(0.2)

    if newly_processed:
        state["processed_message_ids"] = list(processed_ids.union(newly_processed))[-1000:]
        save_state(state)

    return {"status": "ok", "processed_messages": len(newly_processed)}

