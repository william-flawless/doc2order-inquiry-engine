from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from src.common.errors import MissingRequiredHeaderError, WorkbookValidationError
from src.common.s3_utils import get_object_bytes_io


EXPECTED_SHEET_NAME = "Sheet1"
HEADER_ROW_INDEX = 1
REQUIRED_HEADERS = ["Sku", "Category", "Brand", "Description", "Qty", "Stock", "Price"]


def load_workbook_from_s3(bucket: str, key: str):
    """
    Load workbook bytes from S3 into openpyxl workbook.
    """
    try:
        file_obj = get_object_bytes_io(bucket, key)
        workbook = load_workbook(file_obj, data_only=True)
        return workbook
    except Exception as exc:
        raise WorkbookValidationError(f"Failed to load workbook from s3://{bucket}/{key}: {exc}") from exc


def get_sheet(workbook, expected_sheet_name: str = EXPECTED_SHEET_NAME):
    """
    Return the expected sheet. Fail fast if missing.
    """
    if expected_sheet_name not in workbook.sheetnames:
        raise WorkbookValidationError(
            f"Expected sheet '{expected_sheet_name}' not found. Available sheets: {workbook.sheetnames}"
        )
    return workbook[expected_sheet_name]


def read_header_map(sheet, header_row_index: int = HEADER_ROW_INDEX) -> dict[str, int]:
    """
    Build a mapping of header name -> 1-based column index.
    """
    header_map: dict[str, int] = {}

    for col_idx, cell in enumerate(sheet[header_row_index], start=1):
        value = cell.value
        if value is None:
            continue

        header = str(value).strip()
        if header:
            header_map[header] = col_idx

    return header_map


def validate_headers(
    header_map: dict[str, int],
    required_headers: list[str] | None = None,
) -> None:
    """
    Validate that all required headers exist exactly.
    """
    required_headers = required_headers or REQUIRED_HEADERS
    missing = [header for header in required_headers if header not in header_map]

    if missing:
        raise MissingRequiredHeaderError(
            f"Missing required headers: {missing}. Found headers: {sorted(header_map.keys())}"
        )


def _cell_value(sheet, row_idx: int, col_idx: int) -> Any:
    return sheet.cell(row=row_idx, column=col_idx).value


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _row_has_data(row_data: dict[str, Any]) -> bool:
    """
    A row counts as a meaningful inquiry row if at least one of the key input fields is populated.
    """
    return not (
        _is_blank(row_data.get("Sku"))
        and _is_blank(row_data.get("Description"))
        and _is_blank(row_data.get("Qty"))
    )


def extract_stock_inquiry_rows(
    sheet,
    header_map: dict[str, int],
    data_start_row: int = HEADER_ROW_INDEX + 1,
    blank_row_stop_threshold: int = 3,
) -> list[dict[str, Any]]:
    """
    Extract Evapo stock inquiry rows from the worksheet.
    """
    extracted_rows: list[dict[str, Any]] = []
    consecutive_blank_rows = 0
    max_row = sheet.max_row

    for row_idx in range(data_start_row, max_row + 1):
        raw_row_data = {
            header: _cell_value(sheet, row_idx, col_idx)
            for header, col_idx in header_map.items()
        }

        if _row_has_data(raw_row_data):
            consecutive_blank_rows = 0
            extracted_rows.append(
                {
                    "row_index": row_idx,
                    "raw_row_data": raw_row_data,
                }
            )
        else:
            consecutive_blank_rows += 1
            if consecutive_blank_rows >= blank_row_stop_threshold:
                break

    return extracted_rows


def parse_evapo_stock_inquiry_workbook(
    bucket: str,
    key: str,
) -> dict[str, Any]:
    """
    Full parser entrypoint for Evapo stock inquiry workbook.
    """
    workbook = load_workbook_from_s3(bucket, key)
    sheet = get_sheet(workbook, EXPECTED_SHEET_NAME)
    header_map = read_header_map(sheet, HEADER_ROW_INDEX)
    validate_headers(header_map, REQUIRED_HEADERS)
    raw_rows = extract_stock_inquiry_rows(sheet, header_map)

    if not raw_rows:
        raise WorkbookValidationError("No valid inquiry rows found in workbook.")

    return {
        "sheet_name": sheet.title,
        "header_map": header_map,
        "raw_rows": raw_rows,
    }